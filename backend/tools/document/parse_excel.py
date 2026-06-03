"""
V16 R1：Excel 文档工具（MCP-compatible Adapter）。

技术路线（按技术方案 CSV 第 6 行）：
  - openpyxl 主方案：读取 sheet、行列、单元格值
  - MarkItDown 预览兜底（仅当 openpyxl 未安装时）
  - 两者都未安装 → error_code=dependency_missing

metadata 至少包含：
  source_type=xlsx / filename / sheet_name / sheet_names /
  row_start / row_end / column_names / table_index /
  merged_cell_policy / parser_name / extract_method / content_hash
"""

from __future__ import annotations

import hashlib
import time
from pathlib import Path
from typing import Any

from tools.document.errors import (
    DEPENDENCY_MISSING,
    EMPTY_SHEET,
    FILE_NOT_FOUND,
    FILE_TOO_LARGE,
    INVALID_EXCEL,
    UNSUPPORTED_FILE_TYPE,
)
from tools.document.limits import (
    MAX_EXCEL_CELLS,
    MAX_EXCEL_ROWS,
    MAX_EXCEL_SHEETS,
    MAX_FILE_BYTES,
)
from tools.document.quality import assess_quality
from tools.document.registry import DocumentToolSchema, register
from tools.document.tool_result import DocumentToolResult

_INPUT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "file_path": {"type": "string"},
        "file_content": {"type": ["object", "null"]},
    },
    "required": ["file_path"],
}
_OUTPUT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "status": {"type": "string"},
        "text": {"type": "string"},
        "metadata": {"type": "object"},
        "structured_data": {"type": "object"},
        "quality": {"type": "object"},
        "error_code": {"type": "string"},
    },
}


def _sheet_to_text_and_meta(ws: Any, max_rows: int = MAX_EXCEL_ROWS) -> tuple[str, dict]:
    """把一个 worksheet 转成文本行 + 元数据。"""
    rows_data: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = [str(c) if c is not None else "" for c in row]
        rows_data.append(cells)

    if not rows_data:
        return "", {"row_count": 0, "col_count": 0, "column_names": [], "is_empty": True}

    # 第一行作为列名
    header = rows_data[0]
    data_rows = rows_data[1:]

    lines = [" | ".join(header)]
    for row in data_rows:
        lines.append(" | ".join(row))

    text = "\n".join(lines)
    return text, {
        "row_count": len(rows_data),
        "col_count": len(header),
        "column_names": header,
        "row_start": 1,
        "row_end": len(rows_data),
        "is_empty": False,
    }


def _parse_excel(
    file_path: str,
    file_content: bytes | None = None,
) -> DocumentToolResult:
    """xlsx 文档解析 MCP-compatible Adapter。"""
    t0 = time.monotonic()
    path = Path(file_path)
    file_name = path.name or "unknown"
    trace: list[str] = [f"v16:parse_excel:start file={file_name}"]

    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return DocumentToolResult(
            tool_name="parse_excel",
            source_type="xlsx",
            status="failed",
            error_code=UNSUPPORTED_FILE_TYPE,
            failure_reason=f"不支持的文件类型 {path.suffix}，parse_excel 只接受 .xlsx/.xlsm",
            next_action_hint="如需处理 .xls 格式，请先转换为 .xlsx",
            trace=trace,
        )

    # 读取内容
    if file_content is not None:
        if len(file_content) > MAX_FILE_BYTES:
            return DocumentToolResult(
                tool_name="parse_excel",
                source_type="xlsx",
                status="failed",
                error_code=FILE_TOO_LARGE,
                failure_reason="文件大小超限",
                trace=trace,
            )
        content_bytes = file_content
    else:
        if not path.exists():
            return DocumentToolResult(
                tool_name="parse_excel",
                source_type="xlsx",
                status="failed",
                error_code=FILE_NOT_FOUND,
                failure_reason=f"文件不存在: {file_path}",
                trace=trace,
            )
        content_bytes = path.read_bytes()
        if len(content_bytes) > MAX_FILE_BYTES:
            return DocumentToolResult(
                tool_name="parse_excel",
                source_type="xlsx",
                status="failed",
                error_code=FILE_TOO_LARGE,
                failure_reason="文件大小超限",
                trace=trace,
            )

    # 尝试 openpyxl
    try:
        import openpyxl
    except ImportError:
        # 兜底 MarkItDown
        try:
            from markitdown import MarkItDown
        except ImportError:
            return DocumentToolResult(
                tool_name="parse_excel",
                source_type="xlsx",
                status="failed",
                error_code=DEPENDENCY_MISSING,
                failure_reason="openpyxl 和 MarkItDown 均未安装",
                next_action_hint="请执行 pip install openpyxl",
                trace=trace,
            )
        # MarkItDown 兜底（不提供 sheet 结构，仅预览文本）
        import io
        import os
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content_bytes)
            tmp_path = tmp.name
        try:
            md = MarkItDown()
            result = md.convert(tmp_path)
            text = (result.text_content or "").strip()
        finally:
            try:  # noqa: SIM105
                os.unlink(tmp_path)
            except OSError:
                pass
        if not text:
            return DocumentToolResult(
                tool_name="parse_excel",
                source_type="xlsx",
                status="failed",
                error_code=EMPTY_SHEET,
                failure_reason="MarkItDown 兜底解析后内容为空",
                trace=trace,
            )
        quality = assess_quality(text, "xlsx")
        return DocumentToolResult(
            tool_name="parse_excel",
            source_type="xlsx",
            status="success",
            text=text,
            metadata={
                "source_type": "xlsx",
                "filename": file_name,
                "parser_name": "parse_excel",
                "extract_method": "markitdown_fallback",
                "content_hash": quality.get("content_hash", ""),
            },
            quality=quality,
            warnings=["使用 MarkItDown 兜底，sheet/row/column 结构可能不完整"],
            duration_ms=(time.monotonic() - t0) * 1000,
            trace=trace,
        )

    # openpyxl 主路径
    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True, read_only=True)
    except (OSError, ValueError, RuntimeError) as e:
        return DocumentToolResult(
            tool_name="parse_excel",
            source_type="xlsx",
            status="failed",
            error_code=INVALID_EXCEL,
            failure_reason=f"openpyxl 无法打开文件: {e}",
            trace=trace,
        )

    sheet_names = wb.sheetnames
    trace.append(f"v16:parse_excel:sheets={sheet_names[:5]}")

    if not sheet_names:
        return DocumentToolResult(
            tool_name="parse_excel",
            source_type="xlsx",
            status="failed",
            error_code=EMPTY_SHEET,
            failure_reason="Excel 文件没有可用的 sheet",
            trace=trace,
        )

    # 按上限截断 sheet 数
    active_sheets = sheet_names[: MAX_EXCEL_SHEETS]
    all_parts: list[str] = []
    sheets_meta: list[dict] = []
    total_cells = 0
    has_content = False
    warnings: list[str] = []

    for table_index, sheet_name in enumerate(active_sheets):
        ws = wb[sheet_name]
        sheet_text, sheet_meta = _sheet_to_text_and_meta(ws)
        cell_count = sheet_meta.get("row_count", 0) * sheet_meta.get("col_count", 1)
        total_cells += cell_count

        if total_cells > MAX_EXCEL_CELLS:
            warnings.append(f"已达到单元格上限 {MAX_EXCEL_CELLS}，后续 sheet 已截断")
            break

        if sheet_meta.get("is_empty"):
            warnings.append(f"sheet '{sheet_name}' 为空")
            sheets_meta.append({
                "table_index": table_index,
                "sheet_name": sheet_name,
                "is_empty": True,
            })
            continue

        has_content = True
        sheets_meta.append({
            "table_index": table_index,
            "sheet_name": sheet_name,
            "row_count": sheet_meta.get("row_count", 0),
            "col_count": sheet_meta.get("col_count", 0),
            "column_names": sheet_meta.get("column_names", []),
            "row_start": sheet_meta.get("row_start", 1),
            "row_end": sheet_meta.get("row_end", 1),
            "is_empty": False,
        })
        all_parts.append(f"[Sheet: {sheet_name}]\n{sheet_text}")

    wb.close()

    if not has_content:
        return DocumentToolResult(
            tool_name="parse_excel",
            source_type="xlsx",
            status="failed",
            error_code=EMPTY_SHEET,
            failure_reason="所有 sheet 均为空",
            warnings=warnings,
            trace=trace,
        )

    text = "\n\n".join(all_parts).strip()
    quality = assess_quality(text, "xlsx")
    content_hash = quality.get("content_hash", hashlib.md5(text.encode()).hexdigest())

    metadata: dict[str, Any] = {
        "source_type": "xlsx",
        "filename": file_name,
        "sheet_names": active_sheets,
        "sheet_name": active_sheets[0] if active_sheets else "",
        "table_index": 0,
        "merged_cell_policy": "read_only_flat",
        "parser_name": "parse_excel",
        "extract_method": "openpyxl",
        "content_hash": content_hash,
    }

    structured_data: dict[str, Any] = {
        "sheets": sheets_meta,
        "sheet_count": len(sheets_meta),
    }

    duration_ms = (time.monotonic() - t0) * 1000
    trace.append(
        f"v16:parse_excel:done sheets={len(sheets_meta)} "
        f"text_length={len(text)} quality={quality.get('quality_level')}"
    )

    return DocumentToolResult(
        tool_name="parse_excel",
        source_type="xlsx",
        status="success",
        text=text,
        structured_data=structured_data,
        metadata=metadata,
        quality=quality,
        warnings=warnings + quality.get("warnings", []),
        duration_ms=duration_ms,
        trace=trace,
    )


register(DocumentToolSchema(
    tool_name="parse_excel",
    description="解析 .xlsx Excel 文件，提取 sheet/row/column，返回 ToolResult",
    input_schema=_INPUT_SCHEMA,
    output_schema=_OUTPUT_SCHEMA,
    call_fn=_parse_excel,
))
